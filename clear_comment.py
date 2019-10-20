import inspect
import pandas as pd
import os
import xlrd
from openpyxl import load_workbook
import xlsxwriter
import re
import time as tm
import numpy as np
import gc
import sys


def get_file_name(file_path):
    '''获取一个目录下的文件地址和文件名

    Args:
        file_path: 文件所在的目录

    Return:
        文件的地址,文件的名字
    '''
    file_dir = []
    # 其中os.path.splitext()函数将路径拆分为文件名+扩展名
    for root, dirs, files in os.walk(file_path):
        for file in files:
            # 针对不同格式的数据要进行判断
            if os.path.splitext(file)[1] == '.xlsx':
                file_dir.append(os.path.join(root, file))
            elif os.path.splitext(file)[1] == '.csv':
                file_dir.append(os.path.join(root, file))
    excel_names = []
    for i, file in zip(range(len(files)), files):
        # 判断不同格式的数据，获取文件名称
        if os.path.splitext(file)[1] == '.xlsx':
            excel_name = files[i][:-5]
            excel_names.append(excel_name)
        elif os.path.splitext(file)[1] == '.csv':
            excel_name = files[i][:-4]
            excel_names.append(excel_name)
    return file_dir, excel_names


def read_data(path, file_name):
    """ 同时读取xlsx和csv文件

    Args：
        path: 文件地址
        file_name: 文件的名字

    return:
        DataFrame
    """
    print(file_name)
    create_vars = locals()
    # 针对不同的格式的数据 要进行判断
    if os.path.splitext(path)[1] == '.xlsx':
        # sheet_name = input('sheet_name:')
        # a = open(path)
        create_vars[file_name] = pd.read_excel(
            path, sheet_name='Sheet1', index=False, encoding='utf_8_sig')
        create_vars[file_name]['文件来源'] = file_name
        print("excel, Success")
    if os.path.splitext(path)[1] == '.csv':
        a = open(path, encoding='utf_8')
        create_vars[file_name] = pd.read_csv(
            a, encoding='utf_8_sig', error_bad_lines=False)
        create_vars[file_name]['文件来源'] = file_name
        print("csv, Success")
    return create_vars[file_name]


def concat_file(path, file_names):
    """ 合并一个目录下的多个文件

    Args:
        path: 文件地址
        file_name: 文件名称

    Return:
        文件合并成功的DataFrame

    """
    create_vars = locals()
    count = 0
    df_name = []
    for i, j in zip(path, file_names):
        create_vars[j] = read_data(i, j)
        df_name.append(create_vars[j])
        count += 1
        print(count)
    # concat可以将一个list中的数据全部合并
    data_concat = pd.concat(df_name, sort=True)
    print("concat over")
    return data_concat


def merge_sheet(file_path):
    ''' 合并一个excel中的所有sheet

    Args:
        file_path: 文件的地址

    Return:
        在原文中生成一个all_data的sheet
    '''
    workbook = xlrd.open_workbook(file_path)
    # 获取workbook中所有的表格
    sheets = workbook.sheet_names()
    print("the number of sheets in the excel is : {}".format(sheets))

    # 遍历所有的sheet
    all_data = pd.DataFrame()
    for i in range(len(sheets)):
        sheet_data = pd.read_excel(
            file_path, sheet_name=i, index=False, enconding='utf_8_sig', header=None)
        sheet_data['类别'] = i+1
        all_data = all_data.append(sheet_data)
    all_data.info()

    # 保存为新的sheet
    # 新建sheet
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book

    # 保存为excel
    all_data.to_excel(excel_writer=writer, sheet_name="all_data", index=False)
    writer.save()
    writer.close()
    print('complete merge')
    return all_data


def clear_word(data, col_name):
    # 去掉//@开头的数据
    # data2 = data[~data[col_name].str.contains(
    #   '^//@', regex=True, flags=re.IGNORECASE,na=False)]
    # a = list(b.index)
    # data2 = data.drop(data.index[a])
    # print(len(data2))
    # 去掉不用的符号
    # pattern = r'\[\w*\]|(?=\W)([^，:。\w\s\"@！#])|@\w*|http:\/\/\w*.cn/\w*'
    # data3 = data2[col_name].replace(pattern, '', regex=True)
    # data2[col_name] = data3

    data2 = data.copy()
    # 去掉语气词
    c = data2[col_name].replace(
        r'哈|啊|啦|呵|呀|哇|喵|哎|唔|唉|嗷|诶|嘿|唔|呜|嘤|嘻', '', regex=True)
    data2[col_name] = c
    # 去重
    # data2 = data2.sort_values(by='评论个数', ascending=True)
    # data2 = data2.drop_duplicates(subset=col_name).copy()
    # print('after cleaning length:{}'.format(len(data2)))
    print('去掉了语气词')
    return data2


def add_sheet(dataframe, outfile, name):
    '''

    :param dataframe: 数据
    :param outfile: 保存位置
    :param name: sheet名称
    :return:
    '''
    writer = pd.ExcelWriter(outfile, engine='xlsxwriter')
    if os.path.exists(outfile) != True:
        dataframe.to_excel(writer, name, index=False)
    else:
        book = load_workbook(writer.path)
        writer.book = book
        dataframe.to_excel(engine="xlsxwriter", sheet_name=name,
                           index=False, excel_writer='xlsxwriter')
    writer.save()
    writer.close()


def count_kw(kw_list, data, column):
    ''' 统计一个关键词list的讨论量

    Args:
        kw_list: 关键词, list, series
        data: 数据源，dataframe
        columns: 在哪个字段(列名)内查找频次,string

    Return: 
        返回一个关键词及次数的dataframe
    '''
    value = []
    row = []
    for i in kw_list.dropna():
        a = len(data[data[column].str.contains(
            str(i), regex=True, flags=re.IGNORECASE, na=False)])
        value.append(a)
        row.append(i)
        print([i, a])
    dic = {'关键词': row, '次数': value}
    df = pd.DataFrame(dic)
    return df


def str_2_date(x):
    """ 待改善的日期调整

    """
    if re.findall(r'\d{1,4}', str(x)):
        num = re.findall(r'\d{1,4}', str(x))
        date_str = '/'.join(num)
        date = tm.strptime(date_str, '%Y/%m/%d')
        return date
    else:
        return np.nan


def cross_analysis(data, columns, token=','):
    ''' 做交叉分析。将一个cell中的值，拆分为多行。

    Args:
        data: 数据,dataframe
        columns: 要处理的列名,string
        token: 分割符号, string

    Retrun: 
        拆分好的数据框。因为是将同一个cell中的数据进行拆分，所以要保证一个cell
        拆分后对应的索引不变，这会让我们找到拆分后的数据在原数据中的位置，也方
        便通过统计不同的索引次数，来统计单个值的讨论量
    '''

    a = data[columns].str.split(token, expand=True)
    b = a.stack()    # 压缩成一列：此时的数据结果变为多层列
    # 重新指定index:drop=true，将变更前的列去掉，默认保留
    c = b.reset_index(level=1, drop=True)
    c1 = c.replace(' ', np.nan).dropna()
    e = c1.rename(columns)    # 为series设置名称，方便join
    f = data.drop(columns, axis=1)    # 删除原始数据中的被分出的列，使用join合并：join默认使用index合并
    g = f.join(e)
    return g


def count_table(data, kw, column, platform=False):
    """ 统计一个关键词dataframe的次数

    Args:
        data: 要统计词频的数据
        kw: 关键词表
        column: 在哪一列进行查找
    
    Return:
        返回统计好结果的dataframe，如果需要标注平台，则需要将各平台结果concat
    """
    temp_data_list = []
    for j in kw.columns:
        d = count_kw(kw[str(j)], data, column)
        d['维度'] = str(j)
        temp_data_list.append(d)
    temp_concat = pd.concat(temp_data_list)
    if platform != False:
        temp_concat['平台'] = platform
    return temp_concat

# def e_commerce()

# def bbs()


def label_keywords(kw_df, data, column):
    """ 为数据打上关键词标签

    Args:
        kw_df: 关键词表
        data: 需要打标签的数据
        column: 按照哪个字段进行匹配

    Returns：
        返回一个DataFrame。
        在原有的DataFrame上增加列。

    Raises:
        会有Pandas的warning

    """

    for i in kw_df.columns:
        for j in kw_df[str(i)].dropna():
            data[str(i) + '_' + str(j)] = np.nan
            data[str(i) + '_' + str(j)][data[column].str.contains(str(j),
                                                                  regex=True, na=False)] = str(j)
            print(i, j, len(data[data[column].str.contains(str(j),regex=True, na=False)]))
    return data


def extrac_data(data, column, need_pattern):
    """从数中提取需要的数据

    按照关键词提取包含关键词的文本

    Args:
        data: 从data中提取需要的数据
        column： 对data中的column进行匹配
        need_pattern: 要提取的内容的正则表示

    Retruns:
        返回一个DataFrame
    """
    # data[column] = data[column].str.lower()
    extrac_df = data[data[column].str.contains(
        need_pattern, regex=True, flags=re.IGNORECASE, na=False)]
    print('清洗后数据量:{}'.format(len(extrac_df)))
    return extrac_df


def clear_data(data, column, clear_pattern):
    """从数据中去除不需要的数据

    按照关键词去除包含关键词的文本

    Args:
        data: 从data中清洗掉数据
        column： 对data中的column进行匹配
        need_pattern: 要清洗的内容的正则表示

    Retruns:
        返回一个DataFrame
    """
    # data[column]= data[column].str.lower()
    clear_data = data[data[column].str.contains(
        clear_pattern, regex=True, flags=re.IGNORECASE, na=False) == False]
    print('清洗后数量: {}'.format(len(clear_data)))
    return clear_data


def get_kw(path, sheet_name='Sheet1'):
    """获取关键词：
    """
    kw = pd.read_excel(path, sheet_name, encoding='utf_8_sig')
    # kw = kw.applymap(lambda x: str(x).lower())
    kw = kw.applymap(lambda x: str(x).strip())
    kw.replace('nan', np.nan, inplace=True)
    return kw

# 时间数据处理


# 获取变量名称


def get_variable_name(variable):
    """ 获取变量名的string

    Args:
        variable: 变量
    
    Retrun:
        变量名字的string
    """
    callers_local_vars = inspect.currentframe().f_back.f_locals.items()
    return [var_name for var_name, var_val in callers_local_vars if var_val is variable]


# 获取变量大小
def get_size_of_variable():
    """ 获取全局变量的大小
    Args:

    Retrun:
        返回全局变量及其大小的dataframe。大小的统计单位为mb
    """
    variable_name = []
    variable_size = []
    size_dict = {}
    for i, n in zip(globals().values(), globals().keys()):
        a = round(sys.getsizeof(i)/(np.power(1024, 2)), 2)
        variable_size.append(a)
        variable_name.append(n)
    size_dict = {'variable_name': variable_name,
                 'variable_size': variable_size}
    size_df = pd.DataFrame(size_dict)
    size_df.sort_values(by='variable_size', ascending=False, inplace=True)
    size_df.head()
    return size_df


def delet_var(data_size=10):
    """ 删除高于一定大小变量
    Args:
        data_size: 变量的大小，统计单位mb

    Return:
        回收内存
    """
    size_df = get_size_of_variable()
    del_list = [x for x in size_df['variable_name']
                [size_df['variable_size'] > data_size]]
    for i in del_list:
        del globals()[str(i)]
        print('delet {} !'.format(str(i)))
    gc.collect()


def label_in_one_col(data, kw_df, new_col='topic', pair_col='word'):
    """ 按照关键词为数据添加一列标签

    Args:
         data: 要标记的数据
         kw_df: 关键词dataframe
         new_col: 在data中创建的列名, 默认topic
         pari_col: 在data中要匹配的列名, 默认 word 

    Return:
        在data上多一列new_col的dataframe
    """
    data[new_col] = np.nan
    for i in kw_df.columns:
        for j in kw_df[str(i)].dropna():
            data[new_col][data[pair_col].str.contains(
                str(j), na=False, flags=re.IGNORECASE)] = str(i)
            print(i, j)
    return data


def column_combine(data, need_list, token=',', new_col='word'):
    """ 合并所需的列名，并用符号分开

    Args:
        data: 要合并列的数据
        need_list: 哪些字段需要合并
        token: 用什么符号分开
        new_col: 字段合并后的名字

    Return:
        返回合并好列的数据
    """
    analysis_col = data[str(need_list[0])]
    data.replace(np.nan, ' ', inplace=True)
    for i in need_list[1:]:
        analysis_col = analysis_col + token + data[str(i)]
        print(i)
    data[new_col] = analysis_col
    return data

def date_str(x):
    """ 待改善的日期调整

    """
    if len(x) > 9:
        date_str = re.findall(r'\d{1,4}(?=[年月日/\s])', str(x))
        date_str_combien = '/'.join(date_str)
        return date_str_combien
    else:
        return x

#def jieba:
#segments = []
 #   for row in rows:
  #      content = str(row)
    # TextRank 关键词抽取，只获取固定词性
   #     words = jieba.analyse.extract_tags(content, topK=50,withWeight=False, allowPOS=('a'))     # tfidf 抽取形容词
    #words = jieba.analyse.extract_tags(content, topK=50,withWeight=False, allowPOS=('d'))     # tfidf 抽取副词
    #words = jieba.analyse.extract_tags(content, topK=20,withWeight=False, allowPOS=('n'))     # tfidf 抽取名词
    #words = jieba.cut(content) 不分各种词类,产出所有分词结果
    #    splitedStr = ''
    #for word in words:
        # 停用词判断，如果当前的关键词不在停用词库中才进行记录
            # 记录全局分词
     #   segments.append({'word': word, 'count': 1})
      #  splitedStr += word + ' '
       # dfSg = pd.DataFrame(segments)

    #dfWord = dfSg.groupby('word')['count'].sum()

    #dfWord.to_csv('D:\\Elixir怡丽丝尔\\jieba分词结果.csv', encoding='utf_8_sig') # 导出分词结果